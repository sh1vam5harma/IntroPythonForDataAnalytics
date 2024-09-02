import os
import requests
import json
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, PieChart, Reference, ScatterChart, Series
from dotenv import load_dotenv
from scipy.stats import pearsonr

# Load environment variables from .env file
load_dotenv()

# Retrieve API keys from environment variables
bls_api_key = os.getenv("BLS_API_KEY")
census_api_key = os.getenv("CENSUS_API_KEY")

# BLS API Configuration
bls_series_ids = {
    "national_unemployment_rate": "LNU04000000",  # National unemployment rate
    "veteran_unemployment_rate": "LNU04049526"    # Veteran unemployment rate
}
bls_start_year = "2021"
bls_end_year = "2021"

# Census API Configuration
census_base_url = "https://api.census.gov/data/2021/acs/acs5"
census_params = {
    "get": "B21005_001E,B21005_002E,B21005_003E,B21005_007E,B21001_002E,B21001_003E,B21001_023E,B21001_024E,"
           "B21005_010E,B21005_012E,B21005_011E,B21005_009E,B23003_014E",
    "for": "us:*",  # National data
    "key": census_api_key
}

# Mapping of Census variables to their definitions
census_variable_map = {
    "B21005_001E": "Total Civilian Population 18 Years and Over",
    "B21005_002E": "Civilian Population 18 Years and Over, Male",
    "B21005_003E": "Total Military Population 18 Years and Over",
    "B21005_007E": "Unemployed Veterans, Civilian Population 18 Years and Over",
    "B21001_002E": "Total Male Population 18 Years and Over",
    "B21001_003E": "Total Female Population 18 Years and Over",
    "B21001_023E": "Total Male Veterans 18 Years and Over",
    "B21001_024E": "Total Female Veterans 18 Years and Over",
    "B21005_010E": "Nonveteran, 18 to 34 years, Unemployed",
    "B21005_012E": "Nonveteran, 18 to 34 years, Not in Labor Force",
    "B21005_011E": "Nonveteran, 18 to 34 years, Employed",
    "B21005_009E": "Nonveteran, 18 to 34 years, Total",
    "B23003_014E": "Presence of Own Children Under 18 Years by Age of Householder"
}

def fetch_bls_data():
    """Fetch data from the BLS API."""
    headers = {'Content-type': 'application/json'}
    bls_data = {}

    for label, series_id in bls_series_ids.items():
        data = json.dumps({
            "seriesid": [series_id],
            "startyear": bls_start_year,
            "endyear": bls_end_year,
            "registrationkey": bls_api_key
        })
        response = requests.post('https://api.bls.gov/publicAPI/v2/timeseries/data/', data=data, headers=headers)
        bls_data[label] = json.loads(response.text)
    
    return bls_data

def fetch_census_data():
    """Fetch data from the Census API."""
    response = requests.get(census_base_url, params=census_params)
    return response.json()

def calculate_veteran_unemployment_rate(census_data):
    """Calculate the veteran unemployment rate."""
    total_veteran_population = int(census_data[1][6]) + int(census_data[1][7])  # B21001_023E + B21001_024E
    unemployed_veterans = int(census_data[1][3])  # B21005_007E: Unemployed Veterans
    veteran_unemployment_rate = (unemployed_veterans / total_veteran_population) * 100
    return veteran_unemployment_rate

def calculate_civilian_unemployment_rate(census_data):
    """Calculate the civilian unemployment rate."""
    total_civilian_population = int(census_data[1][0])  # B21005_001E: Total Civilian Population 18 Years and Over
    unemployed_veterans = int(census_data[1][3])  # B21005_007E: Unemployed Veterans
    civilian_unemployment_rate = (unemployed_veterans / total_civilian_population) * 100
    return civilian_unemployment_rate

def calculate_correlation_analysis(national_rates, veteran_rates):
    """Calculate Pearson correlation between national and veteran unemployment rates."""
    correlation, _ = pearsonr(national_rates, veteran_rates)
    return correlation

# Function to add charts to the Excel workbook
def add_charts(workbook):
    """Add various charts to the Excel workbook for data visualization."""
    # Line Chart: National Unemployment Rate vs Veteran Unemployment Rate
    sheet4 = workbook["Trend Analysis"]
    line_chart = LineChart()
    line_chart.title = "Unemployment Rate Trend Analysis"
    line_chart.x_axis.title = "Month"
    line_chart.y_axis.title = "Unemployment Rate"
    
    data = Reference(sheet4, min_col=2, min_row=1, max_col=3, max_row=len(sheet4['A']))
    categories = Reference(sheet4, min_col=1, min_row=2, max_row=len(sheet4['A']))
    line_chart.add_data(data, titles_from_data=True)
    line_chart.set_categories(categories)
    sheet4.add_chart(line_chart, "E5")

    # Scatter Plot: Correlation Analysis
    scatter_chart = ScatterChart()
    scatter_chart.title = "Correlation Between National and Veteran Unemployment Rates"
    scatter_chart.x_axis.title = "National Unemployment Rate"
    scatter_chart.y_axis.title = "Veteran Unemployment Rate"
    
    xvalues = Reference(sheet4, min_col=2, min_row=2, max_row=len(sheet4['A']))
    yvalues = Reference(sheet4, min_col=3, min_row=2, max_row=len(sheet4['A']))
    series = Series(yvalues, xvalues, title_from_data=True)
    scatter_chart.series.append(series)
    sheet4.add_chart(scatter_chart, "L5")

def write_to_excel(bls_data, census_data, veteran_unemployment_rate, civilian_unemployment_rate, correlation, filename):
    """Write BLS and Census data to an Excel file using openpyxl."""
    # Create a new Excel workbook
    workbook = openpyxl.Workbook()

    # Write BLS National Unemployment Rate
    sheet1 = workbook.active
    sheet1.title = "National Unemployment Rate"
    
    headers = ["Year", "Month", "Unemployment Rate"]
    sheet1.append(headers)

    national_rates = []
    veteran_rates = []
    for item in bls_data['national_unemployment_rate']['Results']['series'][0]['data']:
        year = item['year']
        period_name = item['periodName']
        value = item['value']
        sheet1.append([year, period_name, value])
        national_rates.append(float(value))

    # Auto-size columns based on content
    for col in sheet1.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        adjusted_width = (max_length + 2)
        sheet1.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width

    # Write BLS Veteran Unemployment Rate
    sheet2 = workbook.create_sheet(title="Veteran Unemployment Rate")
    sheet2.append(headers)

    for item in bls_data['veteran_unemployment_rate']['Results']['series'][0]['data']:
        year = item['year']
        period_name = item['periodName']
        value = item['value']
        sheet2.append([year, period_name, value])
        veteran_rates.append(float(value))

    # Auto-size columns based on content
    for col in sheet2.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        adjusted_width = (max_length + 2)
        sheet2.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width

    # Write Census Data on Veteran Employment with clear meanings (Pivoted Table)
    sheet3 = workbook.create_sheet(title="Census Veteran Data (2021)")

    # Add header row
    sheet3.append(["Population", "Amount"])

    # Add each demographic data as a row
    for i, var in enumerate(census_data[0]):
        description = census_variable_map.get(var, var)
        value = census_data[1][i]
        sheet3.append([description, value])

    # Auto-size columns based on content
    for col in sheet3.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        adjusted_width = (max_length + 2)
        sheet3.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width

    # Write Trend Analysis comparing National vs Veteran Unemployment Rates
    sheet4 = workbook.create_sheet(title="Trend Analysis")
    sheet4.append(["Month", "National Unemployment Rate", "Veteran Unemployment Rate"])

    # Assuming the data from BLS for both national and veteran unemployment is synchronized by month
    for i in range(len(national_rates)):
        month = bls_data['national_unemployment_rate']['Results']['series'][0]['data'][i]['periodName']
        national_rate = national_rates[i]
        veteran_rate = veteran_rates[i]
        sheet4.append([month, national_rate, veteran_rate])

    # Auto-size columns based on content
    for col in sheet4.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        adjusted_width = (max_length + 2)
        sheet4.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width
    sheet4.append(["Correlation between National and Veteran Unemployment Rates", correlation])

    # Write Civilian and Veteran Unemployment Rates
    sheet5 = workbook.create_sheet(title="Unemployment Rate Summary")
    sheet5.append(["Category", "Unemployment Rate"])
    sheet5.append(["Civilian Population", civilian_unemployment_rate])
    sheet5.append(["Veteran Population", veteran_unemployment_rate])

    # Auto-size columns based on content
    for sheet in [sheet4, sheet5]:
        for col in sheet.columns:
            max_length = max(len(str(cell.value)) for cell in col)
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width

    # Add charts to the workbook
    add_charts(workbook)

    # Save the workbook to the specified location
    workbook.save(filename)
    print(f"Data has been written to {filename}")

def main():
    print("Fetching data from BLS API...")
    bls_data = fetch_bls_data()

    print("Fetching data from Census API...")
    census_data = fetch_census_data()

    print("Calculating Veteran Unemployment Rate...")
    veteran_unemployment_rate = calculate_veteran_unemployment_rate(census_data)

    print("Calculating Civilian Unemployment Rate...")
    civilian_unemployment_rate = calculate_civilian_unemployment_rate(census_data)

    print("Calculating Correlation Analysis...")
    national_rates = [float(item['value']) for item in bls_data['national_unemployment_rate']['Results']['series'][0]['data']]
    veteran_rates = [float(item['value']) for item in bls_data['veteran_unemployment_rate']['Results']['series'][0]['data']]
    correlation = calculate_correlation_analysis(national_rates, veteran_rates)

    # Specify the full path where the file should be saved
    file_path = "/Users/sh1vam5harma/Documents/Vets in Tech Python Course/new team/ViT project workbook.xlsx"
    
    print("Writing data to Excel...")
    write_to_excel(bls_data, census_data, veteran_unemployment_rate, civilian_unemployment_rate, correlation, file_path)


if __name__ == "__main__":
    main()